#!/usr/bin/env python3
"""
V2N Trafik Yonlendirme — Final Analiz Scripti (v3)
====================================================

CIKTI:
    results/analysis_report.xlsx   — 11 sheet
    results/plots/*.png            — grafikler

KULLANIM:
    cd /home/veins/my_workspace/test_workspace/simulte/simulations/cars
    python3 analyze_results.py
"""

import os
import sys
import xml.etree.ElementTree as ET
from pathlib import Path

import pandas as pd
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font

# ────────────────────────────────────────────────────────────
# AYARLAR
# ────────────────────────────────────────────────────────────
RESULTS_DIR = Path("/home/veins/my_workspace/test_workspace/simulte/simulations/cars/results")
PLOTS_DIR   = RESULTS_DIR / "plots"
XLSX_OUT    = RESULTS_DIR / "analysis_report.xlsx"

SCENARIOS = [
    ("Baseline",  "tripinfo_Baseline.xml",       None,                                0),
    ("Pen5",      "tripinfo_Penetration05.xml",  "vehicle_stats_Penetration05.csv",   5),
    ("Pen10",     "tripinfo_Penetration10.xml",  "vehicle_stats_Penetration10.csv",  10),
    ("Pen15",     "tripinfo_Penetration15.xml",  "vehicle_stats_Penetration15.csv",  15),
    ("Pen20",     "tripinfo_Penetration20.xml",  "vehicle_stats_Penetration20.csv",  20),
]

TOP_N_VEHICLES   = 20
DIST_THRESHOLD_M = 1500   # mesafe filtresi esigi (omnetpp.ini ile ayni olmali)

# ────────────────────────────────────────────────────────────

def parse_tripinfo(path: Path) -> dict:
    if not path.is_file():
        return {}
    out = {}
    for _, e in ET.iterparse(str(path), events=("end",)):
        if e.tag != "tripinfo":
            continue
        vid = e.attrib.get("id", "")
        if not vid:
            e.clear()
            continue
        try:
            out[vid] = {
                "duration":   float(e.attrib.get("duration", 0)),
                "waiting":    float(e.attrib.get("waitingTime", 0)),
                "loss":       float(e.attrib.get("timeLoss", 0)),
                "route":      float(e.attrib.get("routeLength", 0)),
                "departEdge": e.attrib.get("departLane", "").rsplit("_", 1)[0],
                "arrivalEdge":e.attrib.get("arrivalLane", "").rsplit("_", 1)[0],
            }
        except (ValueError, TypeError):
            pass
        e.clear()
    return out


def parse_csv(path: Path) -> pd.DataFrame:
    if not path or not path.is_file():
        return pd.DataFrame()
    try:
        return pd.read_csv(path)
    except Exception as ex:
        print(f"  ⚠️  CSV okuma hatasi {path.name}: {ex}")
        return pd.DataFrame()


def main():
    print(f"\n{'='*70}")
    print("V2N TRAFIK YONLENDIRME ANALIZI (v3)")
    print(f"{'='*70}\n")

    PLOTS_DIR.mkdir(exist_ok=True)

    # ── 1. Dosyalari oku ────────────────────────────────────
    print("[1/10] Dosyalar okunuyor...")
    data = {}
    for label, tripinfo_name, csv_name, pen in SCENARIOS:
        tpath = RESULTS_DIR / tripinfo_name
        cpath = RESULTS_DIR / csv_name if csv_name else None
        trip = parse_tripinfo(tpath)
        csv  = parse_csv(cpath) if cpath else pd.DataFrame()
        data[label] = {"trip": trip, "csv": csv, "pen": pen}
        print(f"  {label:10s}: tripinfo={len(trip):5d} arac, csv={len(csv):5d} satir")

    if not data["Baseline"]["trip"]:
        print("\n⚠️  Baseline tripinfo bulunamadi! Cikiliyor.")
        sys.exit(1)

    base_trip = data["Baseline"]["trip"]

    # ── 2. SHEET 01: Genel ozet ─────────────────────────────
    print("\n[2/10] Genel ozet hesaplaniyor...")
    summary_rows = []
    for label, _, _, pen in SCENARIOS:
        trip = data[label]["trip"]
        if not trip:
            continue
        durs = [v["duration"] for v in trip.values()]
        s = pd.Series(durs)
        summary_rows.append({
            "Senaryo":     label,
            "Pen (%)":     pen,
            "n":           len(durs),
            "ort_sure":    round(s.mean(), 2),
            "median":      round(s.median(), 2),
            "min":         round(s.min(), 2),
            "max":         round(s.max(), 2),
            "stddev":      round(s.std(), 2),
            "ort_bekleme": round(pd.Series([v["waiting"] for v in trip.values()]).mean(), 2),
            "ort_kayip":   round(pd.Series([v["loss"]    for v in trip.values()]).mean(), 2),
            "ort_mesafe":  round(pd.Series([v["route"]   for v in trip.values()]).mean(), 2),
        })
    summary_df = pd.DataFrame(summary_rows)

    base_row = summary_df[summary_df["Senaryo"] == "Baseline"].iloc[0]
    for col in ["ort_sure", "ort_bekleme", "ort_kayip", "ort_mesafe"]:
        summary_df[f"{col}_diff"] = (summary_df[col] - base_row[col]).round(2)
        summary_df[f"{col}_pct"]  = ((summary_df[col] - base_row[col]) / base_row[col] * 100).round(2)

    print(summary_df.to_string(index=False))

    # ── 3. SHEET 02: Per-vehicle pairing ────────────────────
    print("\n[3/10] Per-vehicle eslestirme...")
    pairing_rows = []
    per_veh_data = {}
    for label, _, _, pen in SCENARIOS:
        if label == "Baseline":
            continue
        trip = data[label]["trip"]
        if not trip:
            continue
        common = set(base_trip) & set(trip)
        rows = []
        for vid in common:
            b = base_trip[vid]
            p = trip[vid]
            rows.append({
                "vehicleId":     vid,
                "base_duration": b["duration"],
                "pen_duration":  p["duration"],
                "delta_duration":p["duration"] - b["duration"],
                "base_waiting":  b["waiting"],
                "pen_waiting":   p["waiting"],
                "delta_waiting": p["waiting"] - b["waiting"],
                "base_loss":     b["loss"],
                "pen_loss":      p["loss"],
                "delta_loss":    p["loss"] - b["loss"],
                "base_route":    b["route"],
                "pen_route":     p["route"],
                "delta_route":   p["route"] - b["route"],
            })
        df = pd.DataFrame(rows)
        per_veh_data[label] = df
        if len(df) == 0:
            continue
        better = (df["delta_duration"] < -1).sum()
        worse  = (df["delta_duration"] >  1).sum()
        same   = len(df) - better - worse
        pairing_rows.append({
            "Senaryo": label, "Pen (%)": pen, "ortak_arac": len(df),
            "iyilesti (>1s)": int(better), "kotulesti (>1s)": int(worse),
            "ayni (+/-1s)": int(same),
            "ort_delta_sure": round(df["delta_duration"].mean(), 2),
            "median_delta_sure": round(df["delta_duration"].median(), 2),
            "min_delta": round(df["delta_duration"].min(), 2),
            "max_delta": round(df["delta_duration"].max(), 2),
            "stddev_delta": round(df["delta_duration"].std(), 2),
        })
    pairing_df = pd.DataFrame(pairing_rows)
    print(pairing_df.to_string(index=False))

    # ── 4. SHEET 03: Top araclar ────────────────────────────
    print(f"\n[4/10] Top {TOP_N_VEHICLES} iyilesen/kotulesen...")
    top_rows = []
    for label, df in per_veh_data.items():
        if len(df) == 0:
            continue
        s = df.sort_values("delta_duration")
        for tag, sub in [("EN_IYI", s.head(TOP_N_VEHICLES)),
                         ("EN_KOTU", s.tail(TOP_N_VEHICLES).iloc[::-1])]:
            for _, r in sub.iterrows():
                top_rows.append({
                    "Senaryo": label, "Tip": tag,
                    "vehicleId": r["vehicleId"],
                    "base_dur": round(r["base_duration"],1),
                    "pen_dur":  round(r["pen_duration"],1),
                    "delta":    round(r["delta_duration"],1),
                    "base_route_m": round(r["base_route"],1),
                    "pen_route_m":  round(r["pen_route"],1),
                    "delta_route":  round(r["delta_route"],1),
                })
    top_df = pd.DataFrame(top_rows)

    # ── 5. SHEET 04: LTE'li vs LTE'siz (basit) ──────────────
    print("\n[5/10] LTE'li vs LTE'siz karsilastirmasi...")
    lte_rows = []
    for label, _, csv_name, pen in SCENARIOS:
        if label == "Baseline" or not csv_name:
            continue
        trip = data[label]["trip"]
        csv  = data[label]["csv"]
        if not trip or csv.empty:
            continue
        if "sumoId" in csv.columns:
            csv_ids = set(csv["sumoId"].astype(str))
        else:
            csv_ids = set(csv["vehicleId"].astype(str))
        trip_ids = set(trip.keys())
        overlap = csv_ids & trip_ids
        match_ratio = len(overlap) / max(len(csv_ids), 1)
        if match_ratio < 0.5:
            print(f"  ⚠️  {label}: ID eslesme dusuk (ortak: {len(overlap)}/{len(csv_ids)}). Atlanacak.")
            continue
        equipped   = [trip[v]["duration"] for v in overlap]
        unequipped = [trip[v]["duration"] for v in trip_ids - csv_ids]
        lte_rows.append({
            "Senaryo": label, "Pen (%)": pen,
            "LTE_arac_sayi": len(equipped),
            "LTE_ort_sure":  round(pd.Series(equipped).mean(), 2)   if equipped else 0,
            "LTE_median":    round(pd.Series(equipped).median(), 2) if equipped else 0,
            "Normal_arac_sayi": len(unequipped),
            "Normal_ort_sure":  round(pd.Series(unequipped).mean(), 2)   if unequipped else 0,
            "Normal_median":    round(pd.Series(unequipped).median(), 2) if unequipped else 0,
            "Fark_LTE_vs_Normal": round(
                (pd.Series(equipped).mean()   if equipped   else 0) -
                (pd.Series(unequipped).mean() if unequipped else 0), 2),
        })
    lte_df = pd.DataFrame(lte_rows)
    if not lte_df.empty:
        print(lte_df.to_string(index=False))

    # ── 6. Mesafe segmentasyonu ─────────────────────────────
    print(f"\n[6/10] Mesafe segmentasyonu (esik: {DIST_THRESHOLD_M}m)...")
    seg_rows = []
    for label, df in per_veh_data.items():
        if len(df) == 0:
            continue
        for seg_name, mask in [
            (f"<{DIST_THRESHOLD_M}m (filtre)",     df["base_route"] <  DIST_THRESHOLD_M),
            (f"{DIST_THRESHOLD_M}-3000m (orta)",   (df["base_route"] >= DIST_THRESHOLD_M) & (df["base_route"] < 3000)),
            ("3000-5000m (uzun)",                  (df["base_route"] >= 3000) & (df["base_route"] < 5000)),
            (">5000m (cok uzun)",                  df["base_route"] >= 5000),
        ]:
            sub = df[mask]
            if len(sub) == 0:
                continue
            seg_rows.append({
                "Senaryo": label,
                "Segment": seg_name,
                "arac_sayi": len(sub),
                "iyilesti": int((sub["delta_duration"] < -1).sum()),
                "kotulesti": int((sub["delta_duration"] >  1).sum()),
                "ort_delta_sure": round(sub["delta_duration"].mean(), 2),
                "median_delta": round(sub["delta_duration"].median(), 2),
                "ort_base_sure": round(sub["base_duration"].mean(), 1),
                "ort_pen_sure":  round(sub["pen_duration"].mean(), 1),
            })
    segment_df = pd.DataFrame(seg_rows)
    if not segment_df.empty:
        print("\nPen15 segmentasyonu:")
        print(segment_df[segment_df["Senaryo"]=="Pen15"].to_string(index=False))

    # ── 7. SHEET 08: Vehicle Lifecycle ──────────────────────
    print("\n[7/10] Vehicle lifecycle...")
    csv_lte_sets = {}
    for label, _, _, _ in SCENARIOS:
        if label == "Baseline":
            continue
        csv = data[label]["csv"]
        if csv.empty or "sumoId" not in csv.columns:
            csv_lte_sets[label] = set()
            continue
        eq = csv[csv["isEquipped"] == 1] if "isEquipped" in csv.columns else csv
        csv_lte_sets[label] = set(eq["sumoId"].astype(str))

    lifecycle_rows = []
    for vid in base_trip.keys():
        row = {
            "sumoId": vid,
            "baseline_sure":   round(base_trip[vid]["duration"], 1),
            "baseline_mesafe": round(base_trip[vid]["route"], 1),
        }
        for label, _, _, _ in SCENARIOS:
            if label == "Baseline":
                continue
            trip = data[label]["trip"]
            if vid in trip:
                row[f"{label}_sure"]  = round(trip[vid]["duration"], 1)
                row[f"{label}_delta"] = round(trip[vid]["duration"] - base_trip[vid]["duration"], 1)
            else:
                row[f"{label}_sure"]  = None
                row[f"{label}_delta"] = None
            row[f"{label}_isLTE"] = 1 if vid in csv_lte_sets.get(label, set()) else 0
        lifecycle_rows.append(row)
    lifecycle_df = pd.DataFrame(lifecycle_rows)

    # ── 8. SHEET 09: LTE Delta Ozet ─────────────────────────
    print("\n[8/10] LTE'li vs LTE'siz delta ozeti...")
    lte_summary_rows = []
    for label, _, _, pen in SCENARIOS:
        if label == "Baseline":
            continue
        if f"{label}_delta" not in lifecycle_df.columns:
            continue
        df_v = lifecycle_df.dropna(subset=[f"{label}_delta"])
        if df_v.empty:
            continue
        lte = df_v[df_v[f"{label}_isLTE"] == 1]
        non = df_v[df_v[f"{label}_isLTE"] == 0]
        lte_summary_rows.append({
            "Senaryo": label, "Pen (%)": pen,
            "LTE_arac": len(lte),
            "LTE_ort_delta": round(lte[f"{label}_delta"].mean(), 2)   if len(lte) else None,
            "LTE_median_delta": round(lte[f"{label}_delta"].median(), 2) if len(lte) else None,
            "Normal_arac": len(non),
            "Normal_ort_delta": round(non[f"{label}_delta"].mean(), 2)   if len(non) else None,
            "Normal_median_delta": round(non[f"{label}_delta"].median(), 2) if len(non) else None,
            "Fark_LTE_minus_Normal": round(
                (lte[f"{label}_delta"].mean() if len(lte) else 0) -
                (non[f"{label}_delta"].mean() if len(non) else 0), 2),
        })
        if len(lte) and len(non):
            print(f"  {label}: LTE'li ({len(lte)}) ort={lte[f'{label}_delta'].mean():+.2f}s | "
                  f"Normal ({len(non)}) ort={non[f'{label}_delta'].mean():+.2f}s")
    lte_summary_df = pd.DataFrame(lte_summary_rows)

    # ── 9. Edge analizi ─────────────────────────────────────
    print("\n[9/10] Edge analizi...")
    edge_rows = []
    for label, df in per_veh_data.items():
        if len(df) == 0:
            continue
        df_c = df.copy()
        df_c["departEdge"] = df_c["vehicleId"].map(lambda v: base_trip.get(v, {}).get("departEdge", ""))
        worst = df_c.nlargest(100, "delta_duration")
        ec = worst["departEdge"].value_counts().head(15)
        for edge, cnt in ec.items():
            if not edge:
                continue
            edge_rows.append({
                "Senaryo": label, "departEdge": edge,
                "kotulesen_arac": int(cnt),
                "ort_delta": round(worst[worst["departEdge"]==edge]["delta_duration"].mean(), 1),
            })
    edge_df = pd.DataFrame(edge_rows)

    # ── 10. GRAFIKLER ───────────────────────────────────────
    print("\n[10/10] Grafikler ciziliyor...")
    plot_paths = {}
    colors = ["#444", "#1f77b4", "#2ca02c", "#ff7f0e", "#d62728"]

    fig, axes = plt.subplots(2, 2, figsize=(12, 8))
    pens = summary_df["Pen (%)"].tolist()
    for ax, col, title in [
        (axes[0,0], "ort_sure",     "Ortalama Sure (s)"),
        (axes[0,1], "ort_bekleme",  "Ortalama Bekleme (s)"),
        (axes[1,0], "ort_kayip",    "Ortalama Kayip (s)"),
        (axes[1,1], "ort_mesafe",   "Ortalama Mesafe (m)"),
    ]:
        ax.plot(pens, summary_df[col], "o-", linewidth=2, markersize=8)
        ax.set_xlabel("Penetrasyon (%)"); ax.set_ylabel(title); ax.set_title(title)
        ax.grid(True, alpha=0.3)
        for x, y in zip(pens, summary_df[col]):
            ax.annotate(f"{y:.1f}", (x, y), textcoords="offset points",
                        xytext=(0,10), ha="center", fontsize=8)
    fig.suptitle("Penetrasyon Trend Analizi", fontsize=14, fontweight="bold")
    fig.tight_layout()
    p = PLOTS_DIR / "01_trend.png"
    fig.savefig(p, dpi=120, bbox_inches="tight"); plt.close(fig)
    plot_paths["01_trend"] = p

    fig, ax = plt.subplots(figsize=(12, 6))
    for (label, _, _, _), color in zip(SCENARIOS, colors):
        trip = data[label]["trip"]
        if not trip:
            continue
        ax.hist([v["duration"] for v in trip.values()], bins=60, alpha=0.5, label=label, color=color)
    ax.set_xlabel("Sure (s)"); ax.set_ylabel("Arac sayisi")
    ax.set_title("Sure Dagilimi — Tum Senaryolar"); ax.legend(); ax.grid(True, alpha=0.3)
    fig.tight_layout()
    p = PLOTS_DIR / "02_duration_hist.png"
    fig.savefig(p, dpi=120, bbox_inches="tight"); plt.close(fig)
    plot_paths["02_duration_hist"] = p

    n_pens = len(per_veh_data)
    if n_pens > 0:
        fig, axes = plt.subplots(n_pens, 1, figsize=(12, 3*n_pens), squeeze=False)
        for ax, (label, df) in zip(axes.flat, per_veh_data.items()):
            if len(df) == 0:
                continue
            d = df["delta_duration"]
            ax.hist(d, bins=80, color="#1f77b4", alpha=0.7, edgecolor="black")
            ax.axvline(0, color="red", linestyle="--", linewidth=1, label="Degisim yok")
            ax.axvline(d.mean(),   color="green",  linewidth=2, label=f"Ortalama: {d.mean():+.1f}s")
            ax.axvline(d.median(), color="orange", linewidth=2, label=f"Median: {d.median():+.1f}s")
            ax.set_xlabel("Delta (pen - baseline) [s]"); ax.set_ylabel("Arac sayisi")
            ax.set_title(f"{label} — Per-vehicle Delta Dagilimi"); ax.legend(); ax.grid(True, alpha=0.3)
        fig.tight_layout()
        p = PLOTS_DIR / "03_delta_hist.png"
        fig.savefig(p, dpi=120, bbox_inches="tight"); plt.close(fig)
        plot_paths["03_delta_hist"] = p

    if "Pen15" in per_veh_data and len(per_veh_data["Pen15"]):
        df = per_veh_data["Pen15"]
        fig, ax = plt.subplots(figsize=(10, 10))
        cs = ["green" if d < -1 else ("red" if d > 1 else "gray") for d in df["delta_duration"]]
        ax.scatter(df["base_duration"], df["pen_duration"], c=cs, s=8, alpha=0.5)
        lim = max(df["base_duration"].max(), df["pen_duration"].max()) * 1.05
        ax.plot([0, lim], [0, lim], "k--", linewidth=1, label="y=x")
        ax.set_xlabel("Baseline sure (s)"); ax.set_ylabel("Pen15 sure (s)")
        ax.set_title("Baseline vs Pen15 — Per-vehicle\nYesil: iyilesti | Kirmizi: kotulesti")
        ax.set_xlim(0, lim); ax.set_ylim(0, lim); ax.legend(); ax.grid(True, alpha=0.3); ax.set_aspect("equal")
        fig.tight_layout()
        p = PLOTS_DIR / "04_pen15_scatter.png"
        fig.savefig(p, dpi=120, bbox_inches="tight"); plt.close(fig)
        plot_paths["04_pen15_scatter"] = p

    fig, ax = plt.subplots(figsize=(10, 6))
    box_data = []; box_labels = []
    for label, _, _, _ in SCENARIOS:
        trip = data[label]["trip"]
        if not trip:
            continue
        box_data.append([v["duration"] for v in trip.values()])
        box_labels.append(label)
    try:
        bp = ax.boxplot(box_data, tick_labels=box_labels, showfliers=False, patch_artist=True)
    except TypeError:
        bp = ax.boxplot(box_data, labels=box_labels, showfliers=False, patch_artist=True)
    for patch, color in zip(bp["boxes"], colors):
        patch.set_facecolor(color); patch.set_alpha(0.5)
    ax.set_ylabel("Sure (s)"); ax.set_title("Senaryo Bazinda Sure Dagilimi (boxplot)")
    ax.grid(True, alpha=0.3, axis="y")
    fig.tight_layout()
    p = PLOTS_DIR / "05_boxplot.png"
    fig.savefig(p, dpi=120, bbox_inches="tight"); plt.close(fig)
    plot_paths["05_boxplot"] = p

    # G6: Mesafe segmentasyonu - Pen15
    if not segment_df.empty and "Pen15" in segment_df["Senaryo"].values:
        seg15 = segment_df[segment_df["Senaryo"] == "Pen15"]
        fig, ax = plt.subplots(figsize=(11, 5))
        x = range(len(seg15))
        ax.bar(x, seg15["ort_delta_sure"],
               color=["green" if d < 0 else "red" for d in seg15["ort_delta_sure"]])
        ax.set_xticks(x); ax.set_xticklabels(seg15["Segment"], rotation=15)
        ax.set_ylabel("Ortalama Delta (s)")
        ax.set_title("Pen15 — Mesafe Segmentlerine Gore Algoritma Etkisi")
        ax.axhline(0, color="black", linewidth=0.8)
        for i, (d, n) in enumerate(zip(seg15["ort_delta_sure"], seg15["arac_sayi"])):
            ax.text(i, d, f"{d:+.1f}s\n({n} arac)", ha="center",
                    va="bottom" if d >= 0 else "top", fontsize=9)
        ax.grid(True, alpha=0.3, axis="y")
        fig.tight_layout()
        p = PLOTS_DIR / "06_segmentation_pen15.png"
        fig.savefig(p, dpi=120, bbox_inches="tight"); plt.close(fig)
        plot_paths["06_segmentation_pen15"] = p

    print(f"  {len(plot_paths)} grafik {PLOTS_DIR} icine kaydedildi")

    # ── 11. METODOLOJI ─────────────────────────────────────
    methodology = [
        ("01_Genel_Ozet",
         "Her senaryonun TUM araclarinin tripinfo XML'inden okunan sure/bekleme/kayip/mesafe ortalamalari. "
         "diff = senaryo - baseline. Bu sheet tum 6999~ aracin agregat istatistigini verir."),
        ("02_Pairing_Ozet",
         "Baseline ve her Pen senaryosunun ORTAK araclarini (sumoId eslesmesi) eslesti. "
         "Her arac icin delta = pen_sure - baseline_sure hesaplandi. iyilesti=delta<-1s, kotulesti=delta>1s, ayni=arasi."),
        ("03_Top_Araclar",
         f"Her Pen senaryosunda en cok delta'ya sahip {TOP_N_VEHICLES} arac (EN_IYI=en negatif delta, EN_KOTU=en pozitif). "
         "vehicleId ve route uzunlugu degisimi de gosterilir."),
        ("04_LTE_vs_Normal",
         "CSV'de bulunan araclar (LTE'li, isEquipped=1) vs tripinfo'da olup CSV'de olmayanlar (LTE'siz). "
         "Mutlak surelerin ortalamasi karsilastirilir."),
        ("05_PerVeh_<Senaryo>",
         "Her arac icin TAM delta verisi: baseline_sure, pen_sure, delta_sure, ayni 3'lu bekleme/kayip/mesafe icin. "
         "Excel'de filtreleyip kendi analizini yapabilirsin."),
        ("06_Edge_Analiz",
         "Her Pen senaryosunda en cok kotulesen 100 aracin departEdge dagilimi."),
        ("06b_Segmentasyon",
         f"Araclari mesafe segmentlerine bol: <{DIST_THRESHOLD_M}m, {DIST_THRESHOLD_M}-3000m, 3000-5000m, >5000m. "
         "Her segmentin delta'sina bak — algoritma hangi yolculuk tipinde fayda saglar?"),
        ("07_Grafikler",
         "Tum grafikler tek sheet'te gomulu. Daha buyuk halleri results/plots/."),
        ("08_Vehicle_Lifecycle",
         "Her arac icin tum senaryolar yan yana: baseline_sure, Pen5_sure, Pen5_delta, Pen5_isLTE, ... "
         "Belirli bir aracin tum senaryolarda nasil davrandigini gormek icin filtrele (sumoId=X)."),
        ("09_LTE_Delta_Ozet",
         "Her senaryoda LTE'li (isEquipped=1) araclarin ort_delta'si vs LTE'siz araclarin ort_delta'si. "
         "Fark_LTE_minus_Normal: pozitifse LTE'liler kaybediyor, negatifse LTE'liler kazaniyor."),
    ]
    method_df = pd.DataFrame(methodology, columns=["Sheet", "Aciklama"])

    # ── 12. XLSX YAZ ───────────────────────────────────────
    print("\n[XLSX yaziliyor]...")
    with pd.ExcelWriter(XLSX_OUT, engine="openpyxl") as writer:
        summary_df.to_excel(writer, sheet_name="01_Genel_Ozet", index=False)
        if not pairing_df.empty:
            pairing_df.to_excel(writer, sheet_name="02_Pairing_Ozet", index=False)
        if not top_df.empty:
            top_df.to_excel(writer, sheet_name="03_Top_Araclar", index=False)
        if not lte_df.empty:
            lte_df.to_excel(writer, sheet_name="04_LTE_vs_Normal", index=False)
        for label, df in per_veh_data.items():
            if len(df) > 0:
                df.round(2).to_excel(writer, sheet_name=f"05_PerVeh_{label}"[:31], index=False)
        if not edge_df.empty:
            edge_df.to_excel(writer, sheet_name="06_Edge_Analiz", index=False)
        if not segment_df.empty:
            segment_df.to_excel(writer, sheet_name="06b_Segmentasyon", index=False)
        if not lifecycle_df.empty:
            lifecycle_df.to_excel(writer, sheet_name="08_Vehicle_Lifecycle", index=False)
        if not lte_summary_df.empty:
            lte_summary_df.to_excel(writer, sheet_name="09_LTE_Delta_Ozet", index=False)
        method_df.to_excel(writer, sheet_name="10_Metodoloji", index=False)

        wb = writer.book
        ws = wb.create_sheet("07_Grafikler")
        ws["A1"] = "Tum grafikler asagida — daha buyuk halleri plots/ klasorunde"
        ws["A1"].font = Font(bold=True, size=12)
        row = 3
        for name, path in sorted(plot_paths.items()):
            try:
                img = XLImage(str(path))
                img.width = 720; img.height = 480
                ws.add_image(img, f"A{row}")
                row += 28
            except Exception as ex:
                print(f"  ⚠️  Grafik gomulemedi {name}: {ex}")

    print(f"\n{'='*70}")
    print(f"✅ TAMAM!  XLSX: {XLSX_OUT}")
    print(f"           Grafikler: {PLOTS_DIR}/*.png")
    print(f"{'='*70}\n")
    print("Sheet'ler:")
    for s, _ in methodology:
        print(f"  {s}")
    print()


if __name__ == "__main__":
    main()
